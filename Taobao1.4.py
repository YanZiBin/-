#通过手动传输网址，可以爬取浏览页面信息，然后自动点击每个商品进入详细页面，爬取商品详细页面信息

from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
import time
import openpyxl as op
import random

# 全局变量
count = 1  # 写入Excel商品计数

# 启动ChromeDriver服务
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ['enable-automation'])
options.add_argument("--disable-blink-features=AutomationControlled")
driver = webdriver.Chrome(options=options)
driver.maximize_window()
wait = WebDriverWait(driver, 10)


# 定义一个自动下滑的函数
def scroll_down():
    """平滑滚动到页面底部，确保内容完全加载"""
    # 获取初始页面高度
    last_height = driver.execute_script("return document.body.scrollHeight")
    
    while True:
        # 平滑滚动
        for i in range(10):  # 将滚动分成10次执行
            current_height = last_height / 10 * (i + 1)
            driver.execute_script(f"window.scrollTo(0, {current_height});")
            time.sleep(0.5)  # 每次滚动后短暂等待
        
        # 等待页面加载
        time.sleep(2)
        
        # 获取新的页面高度
        new_height = driver.execute_script("return document.body.scrollHeight")
        
        # 如果页面高度没有变化，说明已经到底部
        if new_height == last_height:
            break
            
        last_height = new_height


# 定义函数来爬取主页面的商品信息
def get_goods(total_items):
    global count
    scraped_items = 0
    collected_items = []
    page_number = 1

    while scraped_items < total_items:
        # 平滑滚动页面
        scroll_down()
        
        # 等待商品加载
        try:
            wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.contentInner--xICYBlag > a')))
        except TimeoutException:
            print(f"页面 {page_number} 加载超时")
            break

        html = driver.page_source
        doc = pq(html)
        items = doc('div.contentInner--xICYBlag > a').items()
        
        new_items_found = False
        
        for item in items:
            title = item.find('.title--qJ7Xg_90 span').text()
            price_int = item.find('.priceInt--yqqZMJ5a').text()
            price_float = item.find('.priceFloat--XpixvyQ1').text()
            price = float(f"{price_int}{price_float}") if price_int and price_float else 0.0
            deal = item.find('.realSales--XZJiepmt').text()
            location = item.find('.procity--wlcT2xH9 span').text()
            shop = item.find('.shopNameText--DmtlsDKm').text()
            postText = "包邮" if "包邮" in item.find('.subIconWrapper--Vl8zAdQn').text() else "/"
            t_url = item.attr('href')
            shop_url = item.find('.TextAndPic--grkZAtsC a').attr('href')
            img_url = item.find('.mainPicWrapper--qRLTAeii img').attr('src')

            product = {
                'Page': page_number,
                'Num': count - 1,
                'title': title,
                'price': price,
                'deal': deal,
                'location': location,
                'shop': shop,
                'isPostFree': postText,
                'url': t_url,
                'shop_url': shop_url,
                'img_url': img_url
            }

            # 检查是否为新商品
            if product not in collected_items:
                collected_items.append(product)
                new_items_found = True
                print(f"发现新商品: {title}")

                wb.cell(row=count, column=1, value=1)
                wb.cell(row=count, column=2, value=count - 1)
                wb.cell(row=count, column=3, value=title)
                wb.cell(row=count, column=4, value=price)
                wb.cell(row=count, column=5, value=deal)
                wb.cell(row=count, column=6, value=location)
                wb.cell(row=count, column=7, value=shop)
                wb.cell(row=count, column=8, value=postText)
                wb.cell(row=count, column=9, value=t_url)
                wb.cell(row=count, column=10, value=shop_url)
                wb.cell(row=count, column=11, value=img_url)

                count += 1
                scraped_items += 1

                if scraped_items >= total_items:
                    break

        # 如果没有发现新商品，可能需要翻页或结束
        if not new_items_found:
            print(f"页面 {page_number} 未发现新商品")
            # 尝试点击下一页
            try:
                next_page = driver.find_element(By.CSS_SELECTOR, '下一页按钮的CSS选择器')
                next_page.click()
                page_number += 1
                time.sleep(2)  # 等待新页面加载
            except NoSuchElementException:
                print("没有更多页面了")
                break
        
        print(f"当前已爬取 {scraped_items}/{total_items} 个商品")

    return collected_items


# 定义进入商品详情页面并获取目标信息的函数
# 获取指定详细信息的函数
def simulate_human_scroll():
    """模拟人类的浏览滚动行为"""
    total_height = driver.execute_script("return document.body.scrollHeight")
    current_position = 0
    
    while current_position < total_height:
        # 增加滚动步长以减少总次数(500-700像素)
        scroll_step = random.randint(500, 700)
        current_position += scroll_step
        
        # 滚动到新位置
        driver.execute_script(f"window.scrollTo(0, {current_position});")
        
        # 缩短暂停时间为0.2-0.5秒
        time.sleep(random.uniform(0.2, 0.5))
        
        # 降低停顿概率为10%，停顿时间缩短为0.5-1秒
        if random.random() < 0.1:
            time.sleep(random.uniform(0.5, 1))
            
        # 降低向上滚动概率为5%
        if random.random() < 0.05:
            scroll_up = random.randint(50, 100)
            current_position -= scroll_up
            driver.execute_script(f"window.scrollTo(0, {current_position});")
            time.sleep(random.uniform(0.2, 0.4))

def get_product_detail(url):
    driver.get(url)
    # 缩短初始加载等待时间为0.5-1秒
    time.sleep(random.uniform(0.5, 1))
    
    # 模拟人类浏览行为
    simulate_human_scroll()
    
    detail_texts = []
    
    try:
        detail_elements = driver.find_elements(By.CSS_SELECTOR, "span.valueItemText--HiKnUqGa.f-els-1")
        for element in detail_elements:
            detail_text = element.get_attribute("title") or element.text
            detail_texts.append(detail_text)
    except NoSuchElementException:
        detail_texts.append("N/A")
    
    # 缩短离开页面前的等待时间为0.5-1秒
    time.sleep(random.uniform(0.5, 1))
    
    return ",  ".join(detail_texts)

def Crawer_main(target_url, total_items):
    try:
        driver.get(target_url)
        time.sleep(10)  # 给足够时间加载页面

        # 获取主页面商品列表
        products = get_goods(total_items)

        # 遍历每个商品的链接并进入详细页面抓取
        for idx, product in enumerate(products):
            detail_url = "https:" + product['url']
            detail_texts = get_product_detail(detail_url)
            print(f"第 {idx + 1} 个商品的详细信息：{detail_texts}")

            # 将详细信息保存到 Excel 中
            wb.cell(row=product['Num'] + 1, column=12, value="\n".join(detail_texts))

    except Exception as exc:
        print('Crawer_main函数报错:', exc)

if __name__ == '__main__':
    target_url = input('请输入淘宝搜索结果页面的URL：')
    total_items = int(input('输入要爬取的商品数量：'))

    try:
        ws = op.Workbook()
        wb = ws.active
        wb.append(
            ['Page', 'Num', 'title', 'price', 'deal', 'location', 'shop', 'isPostFree', 'url', 'shop_url', 'img_url',
             'detail_text'])
        Crawer_main(target_url, total_items)
        Filename = '淘宝商品数据.xlsx'
        ws.save(Filename)
        print(Filename + "存储成功~")
    except Exception as exc:
        print("Excel建立失败！", exc)

    # 等待用户手动关闭浏览器
    input("爬取结束，按任意键退出并关闭浏览器...")
